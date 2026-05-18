#!/usr/bin/env python3
"""
Export metrics to different formats for analysis and reporting.

Supported formats:
  - CSV (default)
  - JSON (pretty-printed)
  - HTML (table)
  - Excel (.xlsx)

Usage:
    python export_metrics.py metrics/ --format json --output results.json
    python export_metrics.py metrics/ --format html --output report.html
    python export_metrics.py metrics/ --format excel --output results.xlsx
"""

import csv
import json
import sys
from pathlib import Path
import argparse
from collections import defaultdict


def load_metrics(metrics_dir):
    """Load all metrics from CSV files."""
    metrics_dir = Path(metrics_dir)
    all_data = []
    
    for csv_file in sorted(metrics_dir.glob("metrics_*.csv")):
        with open(csv_file, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not row.get('src_mac'):
                    continue
                # Convert numeric fields
                try:
                    row['throughput_mbps'] = float(row.get('throughput_mbps', 0))
                    row['latency_ms'] = float(row.get('latency_ms', 0))
                    row['jains_fairness_index'] = float(row.get('jains_fairness_index', 0))
                    row['active_paths'] = int(row.get('active_paths', 1))
                    row['total_packets'] = int(row.get('total_packets', 0))
                    row['total_bytes'] = int(row.get('total_bytes', 0))
                    row['multipath_enabled'] = row.get('multipath_enabled', 'False').lower() == 'true'
                    all_data.append(row)
                except (ValueError, TypeError):
                    continue
    
    return all_data


def export_json(data, output_file):
    """Export to JSON format."""
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"[EXPORT] JSON: {output_file}")


def export_csv(data, output_file):
    """Export to CSV format (consolidate all)."""
    if not data:
        print("[ERROR] No data to export")
        return
    
    with open(output_file, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    print(f"[EXPORT] CSV: {output_file}")


def export_html(data, output_file):
    """Export to HTML table format."""
    if not data:
        print("[ERROR] No data to export")
        return
    
    # Group by algorithm
    by_algo = defaultdict(list)
    for row in data:
        by_algo[row['algorithm']].append(row)
    
    html = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Metrics Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { color: #333; }
        h2 { color: #666; margin-top: 30px; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th { background-color: #4CAF50; color: white; padding: 12px; text-align: left; }
        td { border: 1px solid #ddd; padding: 10px; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        tr:hover { background-color: #f0f0f0; }
        .single-path { background-color: #e8f4f8; }
        .multipath { background-color: #e8f8e8; }
    </style>
</head>
<body>
"""
    
    html += f"<h1>Metrics Analysis Report</h1>\n"
    html += f"<p>Generated from {len(data)} flow records</p>\n"
    
    for algorithm in sorted(by_algo.keys()):
        html += f"<h2>{algorithm.upper()}</h2>\n"
        
        by_mode = defaultdict(list)
        for row in by_algo[algorithm]:
            mode = "Multipath" if row['multipath_enabled'] else "Single-Path"
            by_mode[mode].append(row)
        
        for mode in sorted(by_mode.keys()):
            rows = by_mode[mode]
            css_class = "multipath" if mode == "Multipath" else "single-path"
            
            html += f"<h3>{mode} ({len(rows)} flows)</h3>\n"
            html += '<table>\n<tr>\n'
            
            # Table header
            fields = ['flow_id', 'src_mac', 'dst_mac', 'active_paths', 'jains_fairness_index', 'total_bytes', 'total_packets']
            for field in fields:
                html += f"<th>{field}</th>\n"
            html += "</tr>\n"
            
            # Table rows
            for row in rows:
                html += f'<tr class="{css_class}">\n'
                for field in fields:
                    value = row.get(field, '')
                    if isinstance(value, float):
                        value = f"{value:.4f}"
                    elif isinstance(value, int) and field != 'flow_id':
                        value = f"{value:,}"
                    html += f"<td>{value}</td>\n"
                html += "</tr>\n"
            
            html += "</table>\n"
    
    html += """</body>
</html>"""
    
    with open(output_file, 'w') as f:
        f.write(html)
    print(f"[EXPORT] HTML: {output_file}")


def export_excel(data, output_file):
    """Export to Excel format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl not available. Install with: pip install openpyxl")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics"
    
    # Header
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
    # Data rows
    for row_idx, row_data in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = row_data.get(header, '')
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Right-align numbers
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    wb.save(output_file)
    print(f"[EXPORT] Excel: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Export metrics to different formats')
    parser.add_argument('metrics_dir', help='Metrics directory')
    parser.add_argument('--format', '-f', choices=['csv', 'json', 'html', 'excel'], 
                        default='json', help='Export format (default: json)')
    parser.add_argument('--output', '-o', help='Output file (auto-generated if not specified)')
    
    args = parser.parse_args()
    
    print(f"[LOAD] Reading metrics from {args.metrics_dir}...")
    data = load_metrics(args.metrics_dir)
    
    if not data:
        print("[ERROR] No metrics found")
        return 1
    
    print(f"[LOAD] Loaded {len(data)} records")
    
    # Generate output filename if not specified
    if not args.output:
        args.output = f"metrics_export.{args.format}"
        if args.format == 'excel':
            args.output = 'metrics_export.xlsx'
    
    # Export
    if args.format == 'json':
        export_json(data, args.output)
    elif args.format == 'csv':
        export_csv(data, args.output)
    elif args.format == 'html':
        export_html(data, args.output)
    elif args.format == 'excel':
        export_excel(data, args.output)
    
    print(f"[DONE] Export complete!")
    return 0


if __name__ == '__main__':
    sys.exit(main())
